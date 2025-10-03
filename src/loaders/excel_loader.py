"""Excel file loader for XLS and XLSX formats."""

from pathlib import Path
from typing import Iterator, List, Any
import pandas as pd
from openpyxl import load_workbook
from .base_loader import BaseFileLoader


class ExcelLoader(BaseFileLoader):
    """Loader for Excel files (.xls, .xlsx)."""

    SUPPORTED_EXTENSIONS = {".xls", ".xlsx"}

    def supports_extension(self, extension: str) -> bool:
        """Check if extension is supported."""
        return extension.lower() in self.SUPPORTED_EXTENSIONS

    def read_rows(self, filepath: Path, max_rows: int = 50) -> Iterator[List[Any]]:
        """
        Read Excel file row by row.

        Args:
            filepath: Path to the Excel file
            max_rows: Maximum number of rows to read

        Yields:
            List of values for each row
        """
        self.validate_file(filepath)

        try:
            # Use openpyxl for reading row by row (more memory efficient)
            if filepath.suffix == ".xlsx":
                wb = load_workbook(filepath, read_only=True, data_only=True)
                ws = wb.active

                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= max_rows:
                        break
                    # Convert row tuple to list and handle None values
                    yield list(row)

                wb.close()
            else:
                # For .xls files, fall back to reading a small DataFrame
                df = pd.read_excel(
                    filepath, sheet_name=0, header=None, nrows=max_rows, engine="xlrd"
                )
                for _, row in df.iterrows():
                    yield row.tolist()

        except Exception as e:
            raise ValueError(f"Failed to read Excel file {filepath}: {str(e)}")

    def load_from_row(self, filepath: Path, header_row: int) -> pd.DataFrame:
        """
        Load Excel file starting from a specific header row.

        Args:
            filepath: Path to the Excel file
            header_row: Row index where header is located (0-based)

        Returns:
            DataFrame with proper header and data
        """
        self.validate_file(filepath)

        try:
            # Skip rows before the header, then use the next row as header
            # Note: skiprows=header_row means skip rows 0 to header_row-1,
            # then row header_row becomes the header (header=0)
            if header_row > 0:
                df = pd.read_excel(
                    filepath,
                    sheet_name=0,
                    skiprows=header_row,
                    header=0,
                    engine="openpyxl" if filepath.suffix == ".xlsx" else "xlrd",
                )
            else:
                # If header is at row 0, no need to skip
                df = pd.read_excel(
                    filepath,
                    sheet_name=0,
                    header=0,
                    engine="openpyxl" if filepath.suffix == ".xlsx" else "xlrd",
                )

            return df

        except Exception as e:
            raise ValueError(f"Failed to load Excel file {filepath}: {str(e)}")

    def load(self, filepath: Path) -> pd.DataFrame:
        """
        Load Excel file into DataFrame.

        Reads the first sheet without assuming header location.
        """
        self.validate_file(filepath)

        try:
            # Read without header to get raw data
            df = pd.read_excel(
                filepath,
                sheet_name=0,
                header=None,
                engine="openpyxl" if filepath.suffix == ".xlsx" else "xlrd",
            )

            return df

        except Exception as e:
            raise ValueError(f"Failed to load Excel file {filepath}: {str(e)}")
